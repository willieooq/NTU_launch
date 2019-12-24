from flask import Flask, request, abort

from linebot import (
    LineBotApi, WebhookHandler
)
from linebot.exceptions import (
    InvalidSignatureError
)
from linebot.models import *

import  os,random

from openpyxl import load_workbook

from template import *
#file path
path = os.path.join(
        os.path.dirname(__file__),
        'restaurant_list.xlsx')
wb = load_workbook(path)

app = Flask(__name__)

# line bot key
line_channel_access_token ="OA1RYruyUGUDao3jTt8p4ONaTdoclyZV9k8c0QQGYx1V3Px8YVkegHI/6NOeUbeRujess1YFJxiQnH30MFdKzoDBMsLBeMonz8QPe1rB+L3AJP3JQQgKUWzhcsjM/fcWGskXsfxNmy69RYo83WRtggdB04t89/1O/w1cDnyilFU="
line_channel_secret ='6f5adc5e2b6804907f70bdaa6e91bf7c'
to = 'U43bfea6ae84a4af42ff1ebf3c310229d'

# Channel Access Token
line_bot_api = LineBotApi(line_channel_access_token)
# Channel Secret
handler = WebhookHandler(line_channel_secret)
#note
v = {'1':'A','2':'B','3':'C','4':'D','5':'E','6':'F'}
price = {'<100':'1','100~150':'2','$$':'3'}
#function
#option save
def excel_save(user_id,save,location):
    wu = wb['User']
    columns=tuple(wu.columns)
    column=columns[0]
    for cell in column:
        if cell.value == user_id:
            wu[v[str(cell.column+int(location))]+str(cell.row)] = save
            wb.save("restaurant_list.xlsx")
# Post Request from /callback  
@app.route("/callback", methods=['POST'])
def callback():
    # get X-Line-Signature header value
    signature = request.headers['X-Line-Signature']
    # get request body as text
    body = request.get_data(as_text=True)
    app.logger.info("Request body: " + body)
    # handle webhook body
    try:
        handler.handle(body, signature)
    except InvalidSignatureError:
        abort(400)
    return 'OK'
    
#greet message
@handler.add(FollowEvent)
def handle_follow(event):
    #add user id
    ws = wb['User']
    columns = tuple(ws.columns)
    column = columns[0]
    for cell in column:
        print('here')
        if cell.value ==event.source.user_id:
            print('1')
            break
        else:       
            ws.append([event.source.user_id])
            wb.save(path)
            print('sucess')
    print( 'Follow event=',event.source.user_id)#test 
    line_bot_api.reply_message(event.reply_token, [TemplateSendMessage(alt_text='start',template=in_or_out_template)])

# handle message
@handler.add(MessageEvent, message=(TextMessage, LocationMessage))
def handle_message(event):
    list1=['0']
    item = {'user_id':'user_id','option':'start','area':'no','time':'0','price':'0','variable':'0'}
    #read excel
    wu = wb['User']
    ws = wb['Restaurant']
    #find user id and option
    columns=tuple(wu.columns)
    column=columns[0]
    for cell in column:
        if cell.value ==  event.source.user_id:
            item['user_id'] = cell.value
            item['option'] = wu[v[str(cell.column+1)]+str(cell.row)].value
            item['area'] = wu[v[str(cell.column+2)]+str(cell.row)].value
            item['time'] = wu[v[str(cell.column+3)]+str(cell.row)].value
            item['price'] = wu[v[str(cell.column+4)]+str(cell.row)].value
            item['variable'] = wu[v[str(cell.column+5)]+str(cell.row)].value 
            print('option=',item['option'])#test
            wb.save("restaurant_list.xlsx")
    user_message = event.message.text
    if isinstance(event.message, TextMessage):
        #location
        if (user_message == '校內' and item['option'] == 'start'):
            line_bot_api.push_message(item['user_id'],TemplateSendMessage(alt_text='in',template=school))
            item['option'] = 'area'
            excel_save(item['user_id'],item['option'],1)
        elif (user_message == '校外' and item['option'] == 'start'):
            line_bot_api.push_message(item['user_id'],TemplateSendMessage(alt_text='out',template=out))
            item['option'] = 'area'
            excel_save(item['user_id'],item['option'],1)
        elif((user_message == 'A' or user_message == 'B' or user_message == 'C' or user_message == 'D' or user_message == 'E' or user_message == 'F' ) and item['option'] == 'area'):
            item['area'] = user_message
            item['option'] = 'time'
            print('here')
            excel_save(item['user_id'],item['area'],2)
            excel_save(item['user_id'],item['option'],1)
            line_bot_api.push_message(item['user_id'],TemplateSendMessage(alt_text='time',template=time))
        elif((user_message == '>30'or user_message == '<30' or user_message =="infinite") and item['option'] == 'time'):
            item['time'] = user_message
            item['option'] = 'price'
            excel_save(item['user_id'],item['time'],3)
            excel_save(item['user_id'],item['option'],1)
            line_bot_api.push_message(item['user_id'],TemplateSendMessage(alt_text='time',template=price_tem))
        elif((user_message == '<100' or user_message == '100~150' or user_message == '$$') and item['option'] == 'price')or(user_message == 'restart' and item['option'] == 'result'):
            line_bot_api.push_message(item['user_id'],TextSendMessage(text = '處理中請稍等...'))
            if user_message != 'restart':
                item['price'] = user_message
                item['option'] = 'result'
                excel_save(item['user_id'],item['price'],4)
                excel_save(item['user_id'],item['option'],1)
            #fliter
            columns=tuple(ws.columns)
            column1=columns[1]
            column0=tuple(ws.columns)[0]
            for cell1 in column1:
                for address in location(item['area'],item['time']):
                    if cell1.value ==  address:
                        column2=columns[2]
                        for cell2 in column2:
                            if ws[v[str(cell2.column-2)]+str(cell2.row)].value != None:
                                if cell2.value ==  int(price[item['price']]):
                                    list1.append(ws[v[str(cell2.column-2)]+str(cell2.row)].value)
                                elif (int(price[item['price']]) == 3):
                                    list1.append(ws[v[str(cell2.column-2)]+str(cell2.row)].value)
            item['variable'] = chr(len(list1)-1)
            excel_save(item['user_id'],item['variable'],5)
            restaurant = list1[random.randint(0,len(list1)-1)]
            for cell in column0:
                if cell.value == restaurant:
                    print(list(ws.rows)[cell.row][0].value)
                    line_bot_api.push_message(item['user_id'],[ TextSendMessage(text = list(ws.rows)[cell.row][0].value)
                                                                ,TextSendMessage(text = list(ws.rows)[cell.row][3].value)
                                                                ,TextSendMessage(text = list(ws.rows)[cell.row][4].value)
                                                                ,TemplateSendMessage(alt_text='OK?',template=ok_tem)])      
        elif (user_message == 'Ok' and item['option'] == 'result'):
            item['option'] = 'start'
            excel_save(item['user_id'],item['option'],1)
            line_bot_api.push_message(item['user_id'],[TextSendMessage(text = '非常感謝您今天的使用，祝您用餐愉快!'),
                                                        TemplateSendMessage(alt_text='66666',template=in_or_out_template)])
        else:
            line_bot_api.reply_message(event.reply_token, [TextSendMessage(text='你喜哩輸入蝦咪挖溝?'),
                                                            TextSendMessage(text='請按照正確的流程來喔!'),
                                                            TemplateSendMessage(alt_text='66666',template=in_or_out_template)])
            item['option'] = 'start'
            excel_save(item['user_id'],item['option'],1)


    
if __name__ == "__main__":
    port = int(os.environ.get('PORT', 5000))
    app.run(host='0.0.0.0', port=port)
